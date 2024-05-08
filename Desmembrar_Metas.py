import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog

# Criação de uma janela Tkinter
root = tk.Tk()
root.title('Automatização planilha de Metas')

# Função para selecionar o arquivo
def selecionar_arquivo():
    arquivo = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Arquivos Excel", "*.xlsx")])
    if arquivo:
        processar_arquivo(arquivo)
        root.destroy()
    else:
        print('Nenhum arquivo selecionado.')

# Função principal para processar o arquivo Excel
def processar_arquivo(arquivo):
    # Carregamento dos dados das planilhas originais
    Meta_Original1 = pd.read_excel(arquivo, sheet_name=0, skiprows=1, nrows=16, usecols='A:H, J:L')
    Meta_Original2 = pd.read_excel(arquivo, sheet_name=1, skiprows=2, nrows=17, usecols='A:AI')
    # Agrupamento dos dados da folha 2 pela coluna 'Cód. Loja'
    grupos_loja = Meta_Original2.groupby('Cód. Loja')
    # Iteração sobre os grupos e criação de uma planilha para cada loja (folha 2)
    for codigo_loja, dados_loja in grupos_loja:
        criar_planilha_para_loja(Meta_Original1, dados_loja, codigo_loja)
    print('Processamento concluído. Arquivos de planilha criados.')

# Função para estilizar o cabeçalho
def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

# Função para criar uma planilha com os dados e estilizar o cabeçalho
def criar_planilha_para_loja(Meta_Geral, Meta_Diarizada, codigo_loja):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = f'Meta_Loja_{codigo_loja}'
    
    # Adiciona cabeçalho da folha 1
    ws1.append(list(Meta_Geral.columns))
    style_header(ws1)
    for _, row in Meta_Geral.iterrows():
        if row['Cód. NOVO'] == codigo_loja:
            ws1.append(row.tolist())  # Adiciona dados da folha 1
            ws1['G2'].number_format = 'R$ ###,###' #Meta Receita Líquida
            ws1['I2'].number_format = '#.##' #Itens/Boleto
            ws1['J2'].number_format = '##.##' #Preço Médio
            ws1['K2'].number_format = '###.##' #Boleto Médio

    # Adiciona uma linha em branco
    ws1.append([])

    # Adiciona o cabeçalho da folha 2 na vertical
    for col_name in Meta_Diarizada.columns:
        ws1.append([col_name, Meta_Diarizada.iloc[0][col_name]])

    # Adiciona os dados da folha 2 na vertical
    for _, row in Meta_Diarizada.iterrows():
        if row['Cód. Loja'] == codigo_loja:
            row_data = []
            for col_name, value in row.items():
                row_data.append(value)  # Adiciona o valor da linha 5
                row_data.append("")  # Adiciona um espaço em branco
    
    # Formatando a meta diarizada em Real (R$)
    cells = ws1['B7':'B38']
    for row in cells:
        for cell in row:
            cell.number_format = 'R$ ##,###'
    
    # Define a largura das colunas de A a L
    for col_idx in range(1, 13):  # Colunas de A a L
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = 23  # Define a largura desejada

    # Centraliza todas as células
    for row in ws1.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Salva a planilha
    wb.save(f'Meta_Loja_{codigo_loja}.xlsx')

# Chamada da função para selecionar o arquivo
selecionar_arquivo()

# Execução do loop principal da janela Tkinter
root.mainloop()
